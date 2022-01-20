# -*- coding:utf-8 -*-
"""
作者：l_jiujiu
日期：2021.11.17
"""
import openpyxl
import numpy as np
import pandas as pd

from plotly.subplots import make_subplots
import plotly.graph_objects as go
import plotly.offline as of  	# 这个为离线模式的导入方法

class Sku:
    def __init__(self, sku_config):
        self.name = sku_config['name']  # sku名称
        self.num = sku_config['num']      # sku序号
        self.sku_time = sku_config['sku_time']  # sku处理所需时间
        self.sku_sectionID=sku_config['sectionID']  # GroupID,sku所在分区

class Section:
    def __init__(self, section_config):
        self.name = section_config['name']  # 分区名称
        self.num = section_config['num']   # 分区编号（main_1为-1，main_2为-2）

        self.waiting_order_list = []   # 分区等待处理订单列表，里面存的是订单实例
        self.process_order_list = []
        self.finish_order_list = []

        self.max_order_num = section_config['max_order_num']  #分区最多积累订单的个数

    # 函数
    def Add_to_waiting_order_list(self, order_now, time):
        # 在section等待订单队列中加入订单
        self.waiting_order_list.append(order_now)
        order_now.now_section_num = self.num
        order_now.time.time_enter_section = time

    def Add_to_finish_order_list(self, order_now):
        # 在section完成订单队列中加入订单
        self.finish_order_list.append(order_now)

    def Count_num(self):
        a = len(self.waiting_order_list) + \
            len(self.process_order_list) + len(self.finish_order_list)
        return a

    def Process_order(self, time):
        # 完成waiting或process中的order：
        if(len(self.process_order_list) == 0):  # process中无：判断waiting中是否有order
            if(len(self.waiting_order_list) == 0):  # waiting中无：return 0
                return 0
            else:  # waiting中有：waiting[0]加入到process中，order_now=process[0]
                self.process_order_list.append(self.waiting_order_list[0])
                self.waiting_order_list.pop(0)
                self.process_order_list[0].time.time_start_process = time

                self.process_order_list[0].time.cal_period_waiting()
                print(
                    '%s' %
                    self.process_order_list[0].name,
                    '在%s' %
                    self.name,
                    '等待用时%d' %
                    self.process_order_list[0].time.period_waiting)
        order_now = self.process_order_list[0]

        # print('工作前%s'%order_now.work_schedule[order_now.now_schedule_num])
        a = int(order_now.work_schedule[order_now.now_schedule_num][1] - 1)
        order_now.work_schedule[order_now.now_schedule_num] = (
            str(self.num), a)
        # print('工作后%s'%order_now.work_schedule[order_now.now_schedule_num])

        # 判断该order_now是否已经完成当前工序，完成则将订单移出process，加入finish
        if(order_now.work_schedule[order_now.now_schedule_num][1] == 0):
            self.finish_order_list.append(self.process_order_list[0])
            self.process_order_list.pop(0)
            # # 测试数据
            # print('%s'%self.name,'中%s'%order_now.name,'已完成任务')
            # print('order.now_schedule_num=%s' % order_now.now_schedule_num, end=',')
            # print('len(order.work_schedule)=%s' % len(order_now.work_schedule))

class Order:
    def __init__(self, order_config):
        self.name = order_config['name']  # 订单名称
        self.num = order_config['num']  # 订单编号

        self.work_schedule = order_config['work_schedule']  # 工序顺序表
        self.now_schedule_num = -1  # 当前工序对应序号
        self.now_section_num = -1  # 当前工序所在section对应序号

        self.time = order_config['time']  # 订单处理实时进度
        self.weighted_cost = 0  #订单cost


    # 函数
    # 计算订单cost
    def Cost_cal(self, section_list):
        j = 0
        cost = 0
        weight = [1, 0.8, 0.5]
        # print(j)
        for i in range(len(self.work_schedule)):
            if (int(self.work_schedule[i][0]) <
                    0):  # 如果第i步为mainstream，跳过这一步判断下一步是否还为mainstream
                continue
            else:
                # cost = 1 *（sec0等待队列 + order在sec0处理时间）+0.8 *（sec1等待队列 + order在sec1处理时间）+0.5 *（sec2等待队列 + order在sec2处理时间）
                try:
                    section_waiting_num = len(section_list[int(self.work_schedule[i][0])].waiting_order_list) + len(section_list[int(
                        self.work_schedule[i][0])].process_order_list) + len(section_list[int(self.work_schedule[i][0])].finish_order_list)
                    cost = cost + \
                        weight[j] * (int(self.work_schedule[i][1]) + section_waiting_num)
                except BaseException:
                    break
                j = j + 1

        self.weighted_cost = cost
        # print("%s" % self.name, "的cost为:%.2f" % self.weighted_cost,"工序为:%s"%self.work_schedule)

class Time:
    def __init__(self, time_config):
        self.order_name = time_config['order_name']  # 所属订单名称
        self.time_enter_section = time_config['time_enter_section']  # 进入分区时间
        self.time_start_process = time_config['time_start_process']  # 开始拣选时间
        self.time_leave_section = time_config['time_leave_section']  # 离开分区时间

        self.period_process = time_config['period_process']  # 拣选共计用时
        self.period_waiting = 0  # 在分区等待共计用时

    # 函数
    # 计算拣选时间
    def cal_period_process(self):
        self.period_process = self.time_leave_section - self.time_start_process + 1

    def cal_period_waiting(self):
        self.period_waiting = self.time_start_process - self.time_enter_section

class Data_Analysis:
    def __init__(self):
        # 画图
        self.x_t = []
        self.y_0 = []
        self.y_0_waiting = []
        self.y_0_process = []

        self.y_1 = []
        self.y_1_waiting = []
        self.y_1_process = []

        self.y_2 = []
        self.y_2_waiting = []
        self.y_2_process = []

        self.y_3 = []
        self.y_3_waiting = []
        self.y_3_process = []

        self.y_4 = []
        self.y_4_waiting = []
        self.y_4_process = []

        self.y_5 = []
        self.y_5_waiting = []
        self.y_5_process = []

        self.y_11 = []
        self.y_22 = []

        self.main_jam_1 = 0  # 主路的拥堵情况
        self.main_jam_2 = 0  # 主路的拥堵情况

    # 存数据
    def save_y_t(self,time,section_list,plot):
        plot.x_t.append(time)
        for i in range(len(section_list)):
            a = section_list[i].Count_num()
            if (section_list[i].num == -1):
                plot.y_11.append(a)
            elif (section_list[i].num == -2):
                plot.y_22.append(a)
            else:
                exec("plot.y_{}.append(a)".format(i))
                exec("plot.y_{}_waiting.append(len(section_list[i].waiting_order_list))".format(i))
                exec("plot.y_{}_process.append(len(section_list[i].process_order_list))".format(i))

    # plotly画折线图
    def plot_scatter(self,y,title,key):
        color = {
            'all': ['circle','blue'],
            'waiting': ['circle-open','orange'],
            'process': ['square','green'],
        }
        section_all = go.Scatter(
            x=self.x_t,
            y=y,
            name=title,  # Style name/legend entry with html tags
            connectgaps=True,
            marker=dict(symbol=color[key][0], color=color[key][1]),
        )
        return section_all

    # plotly画多组折线图
    def plot_results_plotly(self):
        # 用plotly画图
        fig = make_subplots(
            rows=8, cols=1,
            subplot_titles=("<b>1701</b>",
                             "<b>1801</b>",
                             "<b>1901</b>",
                             "<b>2001</b>",
                             "<b>2101</b>",
                             "<b>2201</b>",
                             "<b>主路-1</b>",
                             "<b>主路-2</b>",
                             )
        )
        list_test = [0, 1, -1, 2, 3, -2, 4, 5]
        fig_name=['1701','1801','1901','2001','2101','2201']
        for i in list_test:
            if(i==-1):
                fig.add_trace(self.plot_scatter(self.y_11, '主路<b>%d</b>' % i,key='all'), row = 7, col = 1)
            elif(i==-2):
                fig.add_trace(self.plot_scatter(self.y_22, '主路<b>%d</b>' % i,key='all'), row = 8, col = 1)
            else:

                exec("fig.add_trace(self.plot_scatter(self.y_{},'<b>%d</b> '%{},key='all'),row=i+1,col=1)".format(i,fig_name[i]))
                # exec("fig.add_trace(self.plot_scatter(self.y_{}_waiting, 'section_<b>%d</b> waiting' % {},key='waiting'), row=i + 1, col=1)".format(i,i))
                # exec("fig.add_trace(self.plot_scatter(self.y_{}_process, 'section_<b>%d</b> process' % {},key='process'), row=i + 1, col=1)".format(i,i))

        layout = go.Layout(title='各时段分区订单累积情况',  # 定义生成的plot 的标题
                           xaxis8={'title':'时间'}, # 定义x坐标名称

                           yaxis1= {'title': '订单数', 'range': [-1, 6]},
                           yaxis2={'title': '订单数', 'range': [-1, 6]},
                           yaxis3={'title': '订单数', 'range': [-1, 6]},
                           yaxis4 = {'title': '订单数', 'range': [-1, 6]},
                           yaxis5 = {'title': '订单数', 'range': [-1, 6]},
                           yaxis6 = {'title': '订单数', 'range': [-1, 6]},
                           yaxis7={'title': '订单数', 'range': [-1, 1]},
                           yaxis8={'title': '订单数', 'range': [-1, 1]},
                           )
        fig.update_layout(layout)

        of.plot(fig,filename = 'Simulation_Result.html')

    # 输出.xls表格
    def xls_output(self,order_start,type):
        wb = openpyxl.Workbook()  # 创建Workbook()对象
        sh_name = wb.sheetnames  # 获取所有sheet
        sh = wb[sh_name[0]]
        sh.title = type
        ws=wb[type]
        for order in order_start:
            ws.append([order.name])  # 往文件中写入数据
        wb.save("Results/OrderSequence.xlsx")  # 保存
        print('\n订单派发顺序已输出')

    def init_sku_time(self,path_sku_time_map,num_sku,Mean,StandardDeviation):
        # 生成time随机数
        random_normal = np.random.normal(Mean, StandardDeviation, num_sku)
        sku_time = []
        for rdm in random_normal:
            if (rdm < 1):
                rdm = 1
            sku_time.append(round(rdm, 0))

        wb_file=path_sku_time_map
        data_time = pd.DataFrame(sku_time)
        data_time.columns = ['Time']
        book = openpyxl.load_workbook(wb_file)  # 读取你要写入的workbook
        writer = pd.ExcelWriter(wb_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data_time.to_excel(writer, sheet_name="Part 1", index=False, startcol=4, startrow=0)
        writer.save()

    def init_sku_time_1(self,path_sku_time_map,num_sku):
        # 生成time随机数
        sku_time=[1]*num_sku

        wb_file=path_sku_time_map
        data_time = pd.DataFrame(sku_time)
        data_time.columns = ['Time']
        book = openpyxl.load_workbook(wb_file)  # 读取你要写入的workbook
        writer = pd.ExcelWriter(wb_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data_time.to_excel(writer, sheet_name="Part 1", index=False, startcol=4, startrow=0)
        writer.save()

class CostList:
    def __init__(self, cost_config):
        self.name = cost_config['name'] #所属订单名称
        self.order = cost_config['order']  # 指排序后的顺序，按升序排列
        self.cost = cost_config['cost']  #weighted_cost
        self.orderfororder = cost_config['orderfororder'] # 用于删除Order序列中已被分配的订单，记录了原始Order的顺序
